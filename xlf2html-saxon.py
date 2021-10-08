import sys
import re
import PySimpleGUI as sg
import subprocess
import datetime
from bs4 import BeautifulSoup
import shutil
import openpyxl


def insert_Excel(translatedHtmlFile, checkedHtmlFile, resultsFile):
    # 結果を入れるエクセルを用意
    shutil.copyfile(xlsxTemplate, resultsFile)

    # 翻訳後のhtmlをオープンしてパース
    with open(translatedHtmlFile, encoding='utf-8') as f:
        translated = f.read()
    soupTranslated = BeautifulSoup(translated, 'html.parser')

    translatedList = []
    for t in soupTranslated.find_all('tr'):
        cols = t.find_all('td')
        src = cols[0].get_text()
        target = cols[1].get_text()
        # print(src, target)
        translatedList.append(src + '\t' + target)
    
    # チェック後のhtmlをオープンしてパース
    with open(checkedHtmlFile, encoding='utf-8') as f:
        checked = f.read()
    soupChecked = BeautifulSoup(checked, 'html.parser')

    checkedList = []
    for t in soupChecked.find_all('tr'):
        cols = t.find_all('td')
        src = cols[0].get_text()
        target = cols[1].get_text()
        checkedList.append(src + '\t' + target)

    # Excelを準備
    wb = openpyxl.load_workbook(resultsFile)
    ws = wb['Sheet1']

    # 翻訳後のテキストを入力する
    countT = 2
    for i in translatedList:
        countStr = str(countT)

        src, target =i.split('\t')
        judge = '=IF(B'+countStr+'=E'+countStr+',"-","check!")'

        srcA = 'A' + countStr
        targetB = 'B' + countStr
        judgeC = 'C' + countStr
        
        ws[srcA].value = src
        ws[targetB].value = target
        ws[judgeC].value = judge
        countT += 1
    
    # チェック後のテキストを入力する
    countC = 2
    for i in checkedList:
        src, target =i.split('\t')
        countStr = str(countC)
        srcA = 'D' + countStr
        targetB = 'E' + countStr
        ws[srcA].value = src
        ws[targetB].value = target
        countC += 1

    # Excelを閉じて保存
    wb.close()
    wb.save(resultsFile)


if __name__ == '__main__':
    sg.theme('Dark Blue 3')

    layout = [
        [sg.Text('xlf file(before):', size=(20, 1)), sg.InputText('', enable_events=True,), sg.FilesBrowse('Add', key='-FILES-', file_types=(('xlf file', '*.xlf'),))],
        [sg.Text('xlf file(after):', size=(20, 1)), sg.InputText('', enable_events=True,), sg.FilesBrowse('Add', key='-FILES-', file_types=(('xlf file', '*.xlf'),))],
        [sg.Text('xsl file:', size=(20, 1)), sg.InputText('', enable_events=True,), sg.FilesBrowse('Add', key='-FILES-', file_types=(('xsl file', '*.xsl'),))],
        [sg.Submit(button_text='Run')]
    ]

    window = sg.Window('xlf2html-saxon', layout)

    while True:
        event, values = window.read()

        # ウィンドウの[x]で終了
        if event is None:
            break

        if event == 'Run':
            f_before = values[0]
            f_after = values[1]
            xsl = values[2]

            # 行頭と行末にダブルクォーテーションがあったら削除
            f_before = re.sub('^\"', '', f_before)
            f_before = re.sub('\"$', '', f_before)
            f_after = re.sub('^\"', '', f_after)
            f_after = re.sub('\"$', '', f_after)
            xsl = re.sub('^\"', '', xsl)
            xsl = re.sub('\"$', '', xsl)

            # OutputするHTMLファイル
            f_before_html = re.sub('xlf$', 'html', f_before)
            f_after_html = re.sub('xlf$', 'html', f_after)

            if f_before == '':
                sg.popup('Please specify a xlf (before) file.')
                continue
            elif f_after == '':
                sg.popup('Please specify a xlf (after) file.')
                continue
            elif xsl == '':
                sg.popup('Please specify a xsl file.')

            cmd1 = 'java' + ' -jar' + ' D:\\tool\\saxonb9-1-0-8j\\saxon9.jar' + ' -s:' + f_before + ' -xsl:' + xsl + ' -o:' + f_before_html
            cmd2 = 'java' + ' -jar' + ' D:\\tool\\saxonb9-1-0-8j\\saxon9.jar' + ' -s:' + f_after + ' -xsl:' + xsl + ' -o:' + f_after_html

            res1 = subprocess.check_call(cmd1)
            res2 = subprocess.check_call(cmd2)
            print(res1)
            print(res2)

            xlsxTemplate = "xliff_diff.xlsx"
            todaydetail = datetime.datetime.today()
            datetime = todaydetail.strftime("%Y%m%d%H%M%S")
            resultsFile = datetime + '_' + xlsxTemplate

            insert_Excel(f_before_html, f_after_html, resultsFile)


            sg.popup('Done!')
    window.close()
    sys.exit()
